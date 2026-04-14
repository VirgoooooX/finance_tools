"""进度监控器"""
import sqlite3
import os
from typing import Optional, List, Dict, Any
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from .models import PaymentSummary
from .database import query_payment_summary, query_payment_details


class ProgressMonitor:
    """进度监控器"""
    
    def __init__(self, db_path: str):
        """
        初始化监控器
        
        参数:
            db_path: 数据库路径
        """
        self.db_path = db_path
    
    def get_summary(self, filters: Optional[Dict] = None) -> PaymentSummary:
        """
        获取收款统计摘要
        
        参数:
            filters: 筛选条件
        
        返回:
            PaymentSummary: 统计摘要
        """
        conn = sqlite3.connect(self.db_path)
        try:
            return query_payment_summary(conn, filters)
        finally:
            conn.close()
    
    def get_details(
        self,
        status: Optional[str] = None,
        filters: Optional[Dict] = None
    ) -> List[Dict[str, Any]]:
        """
        获取收款明细列表
        
        参数:
            status: 状态筛选（'unpaid', 'partial', 'paid'）
            filters: 其他筛选条件
        
        返回:
            明细列表
        """
        conn = sqlite3.connect(self.db_path)
        try:
            if status:
                return query_payment_details(conn, status, filters)
            else:
                # 查询所有状态
                all_details = []
                for s in ['unpaid', 'partial', 'paid']:
                    details = query_payment_details(conn, s, filters)
                    all_details.extend(details)
                return all_details
        finally:
            conn.close()
    
    def export_report(
        self,
        output_path: str,
        filters: Optional[Dict] = None
    ) -> str:
        """
        导出 Excel 报告
        
        参数:
            output_path: 输出文件路径
            filters: 筛选条件
        
        返回:
            生成的文件路径
        """
        wb = Workbook()
        
        # 创建统计摘要工作表
        ws_summary = wb.active
        ws_summary.title = "统计摘要"
        
        summary = self.get_summary(filters)
        
        # 标题样式
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True)
        
        # 写入摘要数据
        ws_summary.append(["收款进度统计摘要"])
        ws_summary.merge_cells('A1:B1')
        ws_summary['A1'].font = Font(size=14, bold=True)
        ws_summary['A1'].alignment = Alignment(horizontal='center')
        
        ws_summary.append([])
        ws_summary.append(["指标", "数值"])
        ws_summary['A3'].fill = header_fill
        ws_summary['A3'].font = header_font
        ws_summary['B3'].fill = header_fill
        ws_summary['B3'].font = header_font
        
        ws_summary.append(["总合同数", summary.total_contracts])
        ws_summary.append(["未收款", summary.unpaid_count])
        ws_summary.append(["部分收款", summary.partial_count])
        ws_summary.append(["已收款", summary.paid_count])
        ws_summary.append(["合同总金额", f"{summary.total_contract_amount:,.2f}"])
        ws_summary.append(["已收款总额", f"{summary.total_received_amount:,.2f}"])
        ws_summary.append(["整体收款比例", f"{summary.overall_payment_ratio:.2%}"])
        ws_summary.append(["生成时间", summary.generated_at.strftime("%Y-%m-%d %H:%M:%S")])
        
        # 调整列宽
        ws_summary.column_dimensions['A'].width = 20
        ws_summary.column_dimensions['B'].width = 25
        
        # 创建明细列表工作表
        ws_details = wb.create_sheet("明细列表")
        
        # 写入表头
        headers = ["合同编号", "客户名称", "合同金额", "已开票金额", "已收款金额", "收款状态", "收款比例", "更新时间"]
        ws_details.append(headers)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws_details.cell(1, col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # 写入明细数据
        details = self.get_details(filters=filters)
        for detail in details:
            ws_details.append([
                detail['contract_number'],
                detail['customer_name'],
                f"{detail['contract_amount']:,.2f}",
                f"{detail['invoiced_amount']:,.2f}",
                f"{detail['received_amount']:,.2f}",
                detail['status'],
                f"{detail['payment_ratio']:.2%}",
                detail['last_updated']
            ])
        
        # 调整列宽
        for col in ws_details.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_details.column_dimensions[column].width = adjusted_width
        
        # 保存文件
        wb.save(output_path)
        return output_path
